"""Excel export for reproducible local-transfer results."""

from __future__ import annotations

import json
from pathlib import Path
import shutil

import pandas as pd


def _replace_sheet(workbook, name: str):
    if name in workbook.sheetnames:
        del workbook[name]
    return workbook.create_sheet(name)


def _write_frame(sheet, frame: pd.DataFrame) -> None:
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))


def export_transfer_workbook(source_workbook: Path | str, artifact_dir: Path | str, output_path: Path | str | None = None) -> Path:
    """Copy the local source workbook and append transfer-result sheets."""
    from openpyxl import load_workbook

    source, artifacts = Path(source_workbook), Path(artifact_dir)
    output = Path(output_path) if output_path else artifacts / f"{source.stem}_with_transfer_results.xlsx"
    required = {name: artifacts / f"transfer_{name}.csv" for name in ("summary", "fold_metrics", "predictions")}
    missing = [str(path) for path in required.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("Artefak transfer belum lengkap: " + ", ".join(missing))
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    book = load_workbook(output)
    for title, path in (("Transfer Summary", required["summary"]), ("Fold Metrics", required["fold_metrics"]), ("Transfer Predictions", required["predictions"])):
        _write_frame(_replace_sheet(book, title), pd.read_csv(path))
    manifest = _replace_sheet(book, "Transfer Manifest")
    manifest.append(["field", "value"])
    details = json.loads((artifacts / "transfer_manifest.json").read_text(encoding="utf-8"))
    for key, value in details.items():
        manifest.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])
    manifest.freeze_panes = "A2"
    manifest.column_dimensions["A"].width, manifest.column_dimensions["B"].width = 34, 100
    book.save(output)
    return output
